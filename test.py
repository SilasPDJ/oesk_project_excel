from openpyxl.styles import PatternFill
from operator import add
import xlsxwriter
import os
from openpyxl.utils import get_column_letter as gcl
import openpyxl
import pandas as pd


def btf_class():
    from bs4 import BeautifulSoup
    with open('test.html') as res:
        soup = BeautifulSoup(res, 'html.parser')
        tables = [str(table) for table in soup.select('table')]
        return tables


# btf()
mylist = pd.read_html('test.html')
with_class = btf_class()
# print(with_class[0])
# input('notaCancelada' in with_class[0])


client_path = r'C:\Users\Silas\OneDrive\_FISCAL-2021\2021\07-2021\Controlesis Tecnologia e Desenvolvimento de Software LTDA'
df = pd.concat([l for l in mylist])

header = ['Nº NF', 'Data', 'Valor', 'Imposto',
          'CPF/CNPJ tomador']
excel_file = os.path.join(client_path, 'tests.xlsx')

# writter = pd.ExcelWriter(excel_file, header=header, index=False)
with pd.ExcelWriter(excel_file, engine='xlsxwriter') as writer:
    df[3] = [d.replace('R$ ', '').replace(',', '.')for d in df[3]]
    df[3] = pd.to_numeric(df[3])

    df[2] = [d.replace('R$ ', '').replace(',', '.')for d in df[2]]
    df[2] = pd.to_numeric(df[2])

    # df.style =
    # create new formats
    book = writer.book
    contabil_format = book.add_format({'num_format': 44})

    add_soma = f'C{len(df[2])+1}'
    add_soma = df[2].sum()
    # df.loc[2] = add_soma

    # input(add_soma)

    while True:
        try:
            wb = df.to_excel(writer, sheet_name='Sheet1',
                             header=header, index=False)
            break
        except ValueError:
            header.append('-')

    worksheet = writer.sheets['Sheet1']

    worksheet.set_column('C:C', None, contabil_format)
    worksheet.set_column('D:D', None, contabil_format)

wb = openpyxl.load_workbook(excel_file)
wks = wb.worksheets
ws = wb.active

# next line = nxln
# last value line = llv
line = nxln = llv = len(df[2]) + 3
llv -= 2

# filt
ws.auto_filter.ref = f'A1:A{llv}'

ln_ret = line + 1
ln_nao = line + 2

# Total independente E DESCARTÁVEL
ws[f'E{line}'] = f'= SUM(C2:C{llv})'

ws[f'C{line}'] = f'= SUM(C{ln_ret}:C{ln_nao})'
ws[f'A{line}'] = f'Valor total'

formul_ret = f'= SUMPRODUCT(SUBTOTAL(9,OFFSET(C2,ROW(C2:C{llv})-ROW(C2),0)),(D2:D{llv}>0)+0)'
n_retforml = f'= SUMPRODUCT(SUBTOTAL(9,OFFSET(C2,ROW(C2:C{llv})-ROW(C2),0)),(D2:D{llv}=0)+0)'

ws[f'C{ln_ret}'] = formul_ret
ws[f'C{ln_ret}'].number_format = ws['C2'].number_format
ws[f'A{ln_ret}'] = 'RETIDO'

ws[f'C{ln_nao}'] = n_retforml


ws[f'A{ln_nao}'] = 'NÃO RETIDO'

# muda
redFill = PatternFill(start_color='FFFF0000',
                      end_color='FFFF0000',
                      fill_type='solid')
for table, row in zip(with_class, ws['A']):
    if 'notaCancelada' in table:
        row.fill = redFill
# ws['A2'].fill = redFill


wb.save(excel_file)
