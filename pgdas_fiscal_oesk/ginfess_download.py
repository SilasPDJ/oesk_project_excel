# dale
import openpyxl
from default.sets import InitialSetting, compt_to_date_obj
from default.webdriver_utilities.wbs import WDShorcuts
from default.interact import press_keys_b4, press_key_b4
from selenium.webdriver import Chrome

from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.common.exceptions import *
from selenium.webdriver.support.ui import Select


from default.webdriver_utilities.pre_drivers import pgdas_driver, ginfess_driver
from time import sleep
from default.webdriver_utilities.pre_drivers import ginfess_driver
from openpyxl import Workbook
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.utils import get_column_letter as gcl
import pandas as pd
import os


class DownloadGinfessGui(InitialSetting, WDShorcuts):

    # only static methods from JsonDateWithDataImprove

    def __init__(self, *dados, compt,  show_driver=False):
        # TODO: settar ginfess_valores nos nos outros municipios
        # em especial São Paulo
        # driver
        __r_social, __cnpj, _ginfess_cod, link = dados

        self.compt = compt
        # mesma coisa de self.any_to_str, só que ele aceita args desempacotados
        self.client_path = self.files_pathit(__r_social.strip(), self.compt)
        self.ginfess_valores = None

        # Checa se já existe certificado
        if _ginfess_cod.lower() == 'não há':
            # removi o ja_imported
            print(
                f'\033[1;31m o cliente {__r_social} não possui notas\n...(muito bom) O certificado anula o _ja_imported...\033[m')
        elif self.check_done(self.client_path, '.png', startswith=__r_social) and self.check_done(self.client_path, '.csv', 'NFSe'):
            # Checka o certificado ginfess, somente
            if show_driver:
                driver = pgdas_driver
                self.__driver__name = driver.__name__
                self.driver = driver = pgdas_driver(self.client_path)
            else:
                driver = ginfess_driver
                self.__driver__name = driver.__name__
                driver = self.driver = ginfess_driver(self.client_path)

            # driver.maximize_window()
            driver.minimize_window()
            self.driver.get(link)
            if self.driver.title != 'NFS-e':  # and 'tremembe' not in self.driver.current_url:
                # self.driver.quit()
                self.driver.close()
                driver = pgdas_driver
                self.__driver__name = driver.__name__
                self.driver = driver = pgdas_driver(self.client_path)
            self.driver.get(link)
            # escolhe driver
            print('tremembe' not in self.driver.current_url)
            # for
            driver = self.driver
            super().__init__(self.driver)
            # if city in 'ABC':
            if self.driver.title == 'NFS-e':
                # links das cidades

                # driver.maximize_window()
                # #######################################################
                self.ABC_ginfess(__cnpj, _ginfess_cod)
                # #######################################################

                try:
                    # Find existent tags
                    driver.implicitly_wait(5)
                    self.tags_wait('table', 'tbody', 'tr', 'td')

                    print('printscreen aqui')

                    self.download()

                    driver.implicitly_wait(5)

                    # Creation initial
                    excel_file = os.path.join(
                        self.client_path, f'{__r_social[:__r_social.find(" ")]}_{__cnpj}.xlsx')
                    # Aqui
                    self.excel_from_html_above(
                        excel_file, html=self.ginfess_table_valores_html_code())

                except IndexError:
                    print('~' * 30)
                    print('não emitiu nenhuma nota'.upper())
                    print('~' * 30)

                driver.save_screenshot(self.certif_feito(
                    self.client_path, add=f"{__r_social}-ginfessDone"))
                # coloquei tudo no dele

            elif self.driver.current_url == 'https://tremembe.sigiss.com.br/tremembe/contribuinte/login.php':
                driver.implicitly_wait(5)

                zero_um = _ginfess_cod.split('//')

                # ginfess login//senha
                self.tags_wait('html')
                self.tags_wait('body')
                while True:
                    driver.implicitly_wait(5)

                    ccm = driver.find_element(By.ID, 'ccm')
                    senha = driver.find_element(By.ID, 'senha')
                    confirma = driver.find_element(By.ID, 'confirma')
                    ccm.send_keys(zero_um[0])

                    for el in ccm, senha, confirma:
                        el.clear()
                    ccm.send_keys(zero_um[0])
                    senha.send_keys(zero_um[1])
                    trem_cod = self.captcha_hacking()
                    sleep(5)
                    # confirma precisa mudar
                    # confirma.send_keys(trem_cod)

                    # driver.find_element(By.ID, 'btnOk').click()
                    if 'login.php' in driver.current_url:
                        driver.refresh()
                        driver.implicitly_wait(6)
                    else:
                        break

                print('break')
                driver.implicitly_wait(10)
                driver.execute_script("""function abre_arquivo(onde){
                var iframe = document.getElementById("main");
                iframe.src = onde;
                }
                """)

                driver.execute_script(
                    "abre_arquivo('dmm/_menuPeriodo.php');")

                driver.implicitly_wait(5)

                # self.tag_with_text('td', 'Movimento ').click()

                sleep(5)

                # iframe = driver.find_element(By.ID, 'main')
                # driver.switch_to.frame(iframe)
                driver.find_element(By.NAME, 'btnAlterar').click()
                driver.implicitly_wait(5)
                select_ano = driver.find_element(By.NAME, 'ano')
                select_mes = Select(driver.find_element(By.NAME, 'mes'))

                # handelling select
                compt = self.compt
                mes, ano = compt.split('-')
                select_mes.select_by_value(mes if len(mes) == 2 else f"0{mes}")
                select_ano.clear()
                select_ano.send_keys(ano)

                driver.find_element(By.NAME, 'ano').clear()
                driver.find_element(By.NAME, 'ano').send_keys(ano)
                mes = self.nome_mes(int(mes))

                driver.find_element(By.XPATH,
                                    f"//select[@name='mes']/option[text()='{mes}']").click()
                # driver.find_element(By.NAME, 'ano').send_keys(ano)

                driver.implicitly_wait(5)
                driver.find_element(By.ID, 'btnOk').click()

                driver.implicitly_wait(10)

                # iframe = driver.find_element(By.ID, 'iframe')
                # driver.switch_to.frame(iframe)

                # self.tag_with_text('td', 'Encerramento').click()
                # driver.switch_to.alert().accept()

                # driver.get('../fechamento/prestado.php')
                # driver.get(f'{url}/nfe/nfe_historico_exportacao.php')
                driver.execute_script(
                    "abre_arquivo('nfe/nfe_historico_exportacao.php');")
                driver.implicitly_wait(3)
                self.tags_wait('html')
                self.tags_wait('body')

                iframe = self.webdriverwait_el_by(By.ID, 'main')
                driver.switch_to.frame(iframe)
                self.webdriverwait_el_by(By.ID, 'todos', 20).click()
                driver.find_element(By.ID, 'btnExportar').click()
                driver.switch_to.alert.accept()

                path_zip = self.client_path
                print(f'path_zip-> {path_zip}')
                self.unzip_folder(path_zip)

                driver.switch_to.default_content()
                driver.save_screenshot(self.certif_feito(
                    self.client_path, add=f"{__r_social}-ginfessDone"))

                # ENCERRAR abaixo
                # TODO: testar encerrar o bendito período trememberal
                def tremembeencerrar_mes():
                    driver.execute_script(
                        "abre_arquivo('fechamento/tomado.php');")
                    driver.implicitly_wait(10)
                    try:
                        driver.find_element(By.ID, 'btnSalvar').click()
                        driver.implicitly_wait(5)
                        driver.switch_to.alert().accept()
                        driver.implicitly_wait(5)
                        # driver.back()
                    except (NoSuchElementException, NoAlertPresentException):
                        print('Já encerrado')

            elif self.driver.current_url == 'https://app.siappa.com.br/issqn_itupeva/servlet/com.issqnwebev3v2.login':
                self.driver.find_element(By.ID, 'vUSR_COD').send_keys(__cnpj)
                self.driver.find_element(By.CSS_SELECTOR,
                                         '[type="password"]').send_keys(_ginfess_cod)
                # d = Chrome().
                press_keys_b4('f9')
                driver.save_screenshot(self.certif_feito(
                    self.client_path, add=f"{__r_social}-ginfessDone"))
            elif 'bragancapaulista.giap.com.br' in self.driver.current_url:
                a = __login, __senha = _ginfess_cod.split('//')

                self.driver.find_element(By.ID,
                                         'P101_USERNAME').send_keys(__login)
                self.driver.find_element(By.CSS_SELECTOR,
                                         '[type="password"]').send_keys(str(__senha))
                self.click_ac_elementors(self.tag_with_text('span', 'ENTRAR'))

                # CONSULTAR
                self.driver.implicitly_wait(30)
                self.driver.execute_script(
                    "javascript:apex.submit('EMISSAO NOTA');")
                mes, ano = self.compt.split('-')
                mes = self.nome_mes(int(mes))
                self.driver.find_element(By.XPATH,
                                         f"//select[@name='P26_MES']/option[text()='{mes}']").click()
                try:
                    self.driver.find_element(By.XPATH,
                                             f"//select[@name='P26_ANO']/option[text()='{ano}']").click()
                except NoSuchElementException:
                    print(
                        f'\033[1;31m{__r_social}, site ainda não atualizou para {ano}...\033[m')
                    self.driver.save_screenshot(self.certif_feito(
                        self.client_path, add=f"{__r_social}-ginfessNOTdone-siteDesatualizado"))
                # CONSULTAR
                else:
                    self.driver.execute_script(
                        "apex.submit({request:'P26_BTN_CONSULTAR'});")
                    print('Digite f9 para continuar')
                    press_key_b4('f9')
                    self.driver.save_screenshot(self.certif_feito(
                        self.client_path, add=f"{__r_social}-ginfessDone"))
            elif 'nfe.prefeitura.sp.gov' in self.driver.current_url:
                driver.execute_script("javascript:location.reload();")
                # while 'login.aspx' in self.driver.current_url:
                # TODO acima... para fazser login

                self.send_keys_anywhere(__cnpj)
                self.send_keys_anywhere(Keys.TAB)
                self.send_keys_anywhere(_ginfess_cod)
                self.send_keys_anywhere(Keys.TAB)

                print('Pressione Enter após login - GINFESS')
                press_keys_b4('Enter')
                # driver.find_element(By.NAME, 'ctl00$body$btEntrar').click()
                sleep(2)
                desired_path = "/consultas.aspx"
                new_url = driver.current_url.rsplit("/", 1)[0] + desired_path
                driver.get(new_url)

                _findelementwait = self.webdriverwait_el_by(
                    By.NAME, 'ctl00$body$ddlExercicio')
                # setta o exercício para consulta
                select_ano = Select(_findelementwait)
                select_mes = Select(driver.find_element(
                    By.NAME, 'ctl00$body$ddlMes'))

                date_compt = compt_to_date_obj(self.compt)
                _mes, _ano = date_compt.month, date_compt.year
                select_ano.select_by_value(str(_ano))
                select_mes.select_by_value(str(_mes))

                def download(value, rename=False) -> str:
                    """Faz download das NFs especificadas filtrando pelo value do input

                    Args:
                        value (str): _description_
                        rename (bool, optional): Se True, renomeia o arquivo baixado. Defaults to False.
                    Returns:
                        Caminho do Arquivo csv que acabou de ser baixado
                    """
                    self.webdriverwait_el_by(
                        By.XPATH, f'//input[@value="{value}"]').click()
                    driver.switch_to.window(driver.window_handles[1])
                    driver.find_element(
                        By.XPATH, '//input[@value="Exportar"]').click()
                    sleep(3)
                    most_recent_file = self.sort_files_by_most_recent(self.client_path)[
                        0]
                    if rename:
                        os.rename(most_recent_file, os.path.join(
                            self.client_path, rename))

                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                    return most_recent_file

                # sem filtrar nfs canceladas
                driver.find_element(By.ID, 'ctl00_body_ckNFCancelada').click()

                download('NFS-e RECEBIDAS', rename='TOMADOR-NFSe.csv')
                csv_file_emitidas = download('NFS-e EMITIDAS')
                self.read_notadomilhao_layout(csv_file_emitidas)
                # ler TODO do backend...
                driver.save_screenshot(self.certif_feito(
                    self.client_path, add=f"{__r_social}-ginfessDone"))
            # elif self.driver.current_url == 'https://sumare.sigissweb.com/':
            else:
                print(__r_social)
                driver.execute_script("javascript:location.reload();")
                self.send_keys_anywhere(__cnpj)
                self.send_keys_anywhere(Keys.TAB)
                self.send_keys_anywhere(_ginfess_cod)
                self.send_keys_anywhere(Keys.TAB)

                if self.__driver__name == "pgdas_driver":
                    from win10toast import ToastNotifier
                    ToastNotifier().show_toast("Pressione F9 para continuar", duration=10)

                    press_keys_b4('f9')
                    driver.save_screenshot(self.certif_feito(
                        self.client_path, add=f"{__r_social}-ginfessDone"))

            [(print(f'Sleeping before close {i}'), sleep(1))
             for i in range(5, -1, -1)]

            driver.close()

    def wait_main_tags(self):
        self.tags_wait('body', 'div', 'table')

    def ABC_ginfess(self, __cnpj, __senha):

        driver = self.driver

        def label_with_text(searched):
            label = driver.find_element(By.XPATH,
                                        f"//label[contains(text(),'{searched.rstrip()}')]")
            return label

        def button_with_text(searched):
            bt = driver.find_element(By.XPATH,
                                     f"//button[contains(text(),'{searched.rstrip()}')]")
            return bt

        def a_with_text(searched):
            link_tag = driver.find_element(By.XPATH,
                                           f"//a[contains(text(),'{searched.rstrip()}')]")
            return link_tag

        self.wait_main_tags()
        """ ~~~~~~~~~~~~~~~~~~~~~~ GLÓRIA A DEUS ~~~~~~~~~~~~~~~~~~"""
        name_c = 'gwt-DialogBox'

        self.del_dialog_box(name_c)

        self.wait_main_tags()
        self.tags_wait('img')
        driver.implicitly_wait(10)

        name_c = 'x-window-dlg', 'ext-el-mask', 'x-shadow'
        try:
            try:
                button_with_text('OK').click()
                for name in name_c:
                    try:  # ...
                        self.del_dialog_box(name)
                    except NoSuchElementException:
                        print('Except dentro do except e no for, [linha 323]')
                        ...
            except (NoSuchElementException, ElementClickInterceptedException):
                pass
            driver.find_element(By.XPATH, '//img[@src="imgs/001.gif"]').click()
        except (NoSuchElementException, ElementClickInterceptedException):
            pass

        try:
            for name in name_c:
                try:  # ...
                    self.del_dialog_box(name)
                except NoSuchElementException:
                    print('Except dentro do except e no for, [linha 336]')
                    ...
                driver.implicitly_wait(5)
            button_with_text('OK').click()
        except (NoSuchElementException, ElementClickInterceptedException):
            print('Sem janela possivel, linha 246')
            # driver.execute_script('window.alert("Não foi possível prosseguir")')

        driver.implicitly_wait(5)
        """ ~~~~~~~~~~~~~~~~~~~~~~ GLÓRIA A DEUS ~~~~~~~~~~~~~~~~~~"""

        # print('mandando teclas...')
        label_with_text("CNPJ:").click()
        self.send_keys_anywhere(__cnpj)
        passwd = driver.find_element(By.XPATH, "//input[@type='password']")

        self.tags_wait('body', 'img')
        passwd.clear()
        passwd.send_keys(__senha)
        button_with_text("Entrar").click()

        # tratando atualiza dados
        driver.implicitly_wait(15)
        try:
            self.wait_main_tags()
            button_with_text('X').click()
            button_with_text('X').click()
            print('CLICADO, X. Linha 263')
        except (NoSuchElementException, ElementClickInterceptedException):
            print('Tentando atualizar os dados')

        a_with_text("Consultar").click()

        print('Waiting main_excel_manager tags')
        self.wait_main_tags()

        period = label_with_text('Período')
        period.click()
        driver.implicitly_wait(5)
        de = label_with_text('De:')
        de.click()
        self.send_keys_anywhere(Keys.BACKSPACE, 10)
        first, last = self.first_and_last_day_compt(
            self.compt, zdate_wontbe_greater=False)
        self.send_keys_anywhere(first)
        driver.implicitly_wait(2.5)
        self.send_keys_anywhere(Keys.TAB)

        self.send_keys_anywhere(last)
        driver.implicitly_wait(5)
        button_with_text("Consultar").click()
        self.wait_main_tags()
        driver.implicitly_wait(10)

    def download(self):
        """
        :city: A, B, C only
        :return:
        """
        driver = self.driver
        try:
            try:
                # self.del_dialog_box('x-shadow')
                xsh = driver.find_element(By.CLASS_NAME, 'x-shadow')
                if 'block' in xsh.get_attribute('style'):
                    self.del_dialog_box('x-shadow')
                    driver.implicitly_wait(10)
                    print('W-SHADOW-DELETADO')
                else:
                    raise NoSuchElementException

            except NoSuchElementException:
                print('Tem notas')
                driver.implicitly_wait(10)
                downloada_xml = driver.find_element(By.XPATH,
                                                    '//img[@src="imgs/download.png"]')
                try:
                    downloada_xml.click()
                except ElementClickInterceptedException:
                    self.click_ac_elementors(downloada_xml)

                except ElementNotInteractableException:
                    pass
                # self.click_ac_elementors(downloada_xml)

        except NoSuchElementException:
            print('NÃO CONSEGUI FAZER DOWNLOAD...')

    def ginfess_table_valores_html_code(self):
        """
        :return: (html_cod): código dele se existe a class ytb-text, scrap_it
        """
        driver = self.driver
        max_value_needed = driver.find_elements(By.CLASS_NAME, 'ytb-text')
        max_value_needed = max_value_needed[1].text[-1]
        print(max_value_needed)
        self.tags_wait('input', 'body')

        cont_export = 1

        xml_pages = driver.find_element(By.XPATH,
                                        '//input[@class="x-tbar-page-number"]')
        driver.implicitly_wait(5)
        number_in_pages = xml_pages.get_attribute('value')

        html_cod = """
                    <style>/*.detalheNota:after{background: red; content: 'cancelada';}*/
                    .notaCancelada{
                    background: red
                    }
                    </style>
                   """.strip()

        for i in range(10):
            print(number_in_pages)
            xml_pages.send_keys(Keys.BACKSPACE)
            xml_pages.send_keys(cont_export)

            xml_pages.send_keys(Keys.ENTER)
            driver.implicitly_wait(5)
            print('CALMA...')
            cont_export += 1
            number_in_pages = xml_pages.get_attribute('value')

            # // div[ @ id = 'a'] // a[ @class ='click']

            wanted_wanted = driver.find_elements(By.XPATH,
                                                 "//div[contains(@class, 'x-grid3-row')]")
            print(wanted_wanted[0].text)
            # table = wanted_wanted

            for w in wanted_wanted:
                # w.click()
                print(w.text)
                html_cod += w.get_attribute('innerHTML')
                # sleep(2)
            # XML_to_excel.ginfess_scrap()
            if int(cont_export) == int(max_value_needed) + 1:
                break
            print('breakou')

            print('~~~~~~~~')
        return html_cod
        # de.send_keys(Keys.TAB)

    def read_notadomilhao_layout(self, plan) -> None:
        """Implementação, gerar nota do milhão automaticamente

        Args:
            plan (os.PathLike): Caminho do csv os.PathLike

        Raises:
            ValueError: Se o valor total não for igual às somas dos outros 2 valores...
        """
        df = pd.read_csv(plan, sep=";", encoding="latin1")
        newdf = df[['Valor dos Serviços', 'ISS Retido']]
        newdf.loc[:, 'Valor dos Serviços'] = newdf['Valor dos Serviços'].str.replace(
            '.', '').str.replace(',', '.').astype(float)

        valor_total = newdf.iloc[:-1, 0].sum()

        # Verifica se é igual ao proposto pelo csv...
        valor_total == newdf.iloc[-1, 0]
        assert valor_total == newdf.iloc[-1, 0]

        # Checa se ISS Retido é único, ou seja se ISS == 'N'
        if newdf.iloc[:-1, 1].nunique() == 1:
            valor_n_retido = valor_total
            valor_retido = 0.00
        else:
            raise ValueError(
                "Necessária Implementação de valor retido na prefeitura de sp")

        self.ginfess_valores = valor_n_retido, valor_retido, valor_total
        assert valor_n_retido + valor_retido == valor_total

    def excel_from_html_above(self, excel_file, html):
        import numpy as np
        from bs4 import BeautifulSoup
        mylist = pd.read_html(html)
        soup = BeautifulSoup(html, 'html.parser')
        nfs_html = [str(table) for table in soup.select('table')]
        # for table, row_a, row_c in zip(with_class, ws['A'], ws['C']):
        # if 'notaCancelada' in table:
        header = ['Nº NF', 'Data', 'Valor', 'Imposto',
                  'CPF/CNPJ tomador', 'NF cancelada']
        # df = pd.concat(mylist)
        df = pd.concat([df.iloc[:, :5] for df in mylist], ignore_index=True)

        def to_number(ind):
            df[ind] = [d.replace('R$ ', '') for d in df[ind]]
            df[ind] = [d.replace('.', '') for d in df[ind]]
            df[ind] = [d.replace(',', '.') for d in df[ind]]
            df[ind] = pd.to_numeric(df[ind])
            return df[ind]

        df[2] = to_number(2)
        df[3] = to_number(3)
        df['NF cancelada'] = [
            'notaCancelada' in status_html for status_html in nfs_html]
        df.columns = range(df.shape[1])  # set back to default the columns name
        df.columns = header

        # df['Valor'].sum()
        is_nf_valid = df['NF cancelada'] == False
        # () evaluates both
        valor_nao_retido = df.loc[(df['Imposto']
                                  == 0) & is_nf_valid, 'Valor'].sum()
        valor_retido = df.loc[(df['Imposto']
                               > 0) & is_nf_valid, 'Valor'].sum()
        valor_total = df.loc[is_nf_valid, 'Valor'].sum()
        # valor_total = df.loc[df['NF cancelada'] == False, 'Valor'].sum()
        # total_sem_cancelar = df['Valor'].sum()
        # is_nf_valid
        # 'NF cancelada' == False then 'NF não foi cancelada, soma'
        # sum_ = df.iloc[df['Status'] == False, df.columns.get_loc('Valor')].sum()
        if 'valor_total' not in df.columns:
            df['valor_nao_retido'] = np.nan
            df['valor_retido'] = np.nan
            df['valor_total'] = np.nan
        df.loc[0, 'valor_nao_retido'] = valor_nao_retido
        df.loc[0, 'valor_retido'] = valor_retido
        df.loc[0, 'valor_total'] = valor_total

        self.ginfess_valores = valor_nao_retido, valor_retido, valor_total

        df.to_excel(excel_file, index=False)

    def excel_from_html_above__old(self, excel_file, html):
        from bs4 import BeautifulSoup
        from openpyxl.styles import PatternFill
        # from win32com.client import Dispatch
        from comtypes.client import CreateObject, GetActiveObject
        # desempenho melhor que a de cima
        import pythoncom

        mylist = pd.read_html(html)

        soup = BeautifulSoup(html, 'html.parser')
        with_class = tables = [str(table) for table in soup.select('table')]
        df = pd.concat([l for l in mylist])
        header = ['Nº NF', 'Data', 'Valor', 'Imposto',
                  'CPF/CNPJ tomador']

        def to_number(ind):
            df[ind] = [d.replace('R$ ', '') for d in df[ind]]
            df[ind] = [d.replace('.', '') for d in df[ind]]
            df[ind] = [d.replace(',', '.') for d in df[ind]]
            df[ind] = pd.to_numeric(df[ind])
            return df[ind]
        with pd.ExcelWriter(excel_file, engine='xlsxwriter') as writer:
            df[2] = to_number(2)
            df[3] = to_number(3)
            # df.style =
            # create new formats
            book = writer.book
            contabil_format = book.add_format({'num_format': 44})

            add_soma = f'C{len(df[2])+1}'
            add_soma = df[2].sum()
            # df.loc[2] = add_soma
            # input(add_soma)
            while True:
                try:
                    wb = df.to_excel(writer, sheet_name='Sheet1',
                                     header=header, index=False)
                    break
                except ValueError:
                    header.append('-')

            worksheet = writer.sheets['Sheet1']

            worksheet.set_column('C:C', None, contabil_format)
            worksheet.set_column('D:D', None, contabil_format)

        wb = openpyxl.load_workbook(excel_file)
        wks = wb.worksheets
        ws = wb.active

        # next line = nxln
        # last value line = llv
        line = nxln = llv = len(df[2]) + 3
        llv -= 2

        # filt
        ws.auto_filter.ref = f'A1:F{llv}'

        ln_ret = line + 1
        ln_nao = line + 2

        # Total independente E DESCARTÁVEL
        ws[f'E{line}'] = f'= SUM(C2:C{llv})'

        ws[f'C{line}'] = f'= SUM(C{ln_ret}:C{ln_nao})'
        ws[f'A{line}'] = f'Valor total'

        formul_ret = f'= SUMPRODUCT(SUBTOTAL(9,OFFSET(C2,ROW(C2:C{llv})-ROW(C2),0)),(D2:D{llv}>0)+0)'
        n_retforml = f'= SUMPRODUCT(SUBTOTAL(9,OFFSET(C2,ROW(C2:C{llv})-ROW(C2),0)),(D2:D{llv}=0)+0)'

        # valores
        ws[f'C{ln_ret}'] = formul_ret
        ws[f'A{ln_ret}'] = 'RETIDO'
        ws[f'C{ln_nao}'] = n_retforml
        ws[f'A{ln_nao}'] = 'NÃO RETIDO'
        # formatacoes
        ws[f'C{ln_ret}'].number_format = ws['C2'].number_format
        ws[f'C{ln_nao}'].number_format = ws['C2'].number_format
        ws[f'C{line}'].number_format = ws['C2'].number_format
        # muda
        redFill = PatternFill(start_color='FFFF0000',
                              end_color='FFFF0000',
                              fill_type='solid')
        #  aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa
        with_class.insert(0, '')
        for table, row_a, row_c in zip(with_class, ws['A'], ws['C']):
            if 'notaCancelada' in table:
                row_a.fill = redFill
                row_c.value = ""  # nota_cancelada
        # ws['A2'].fill = redFill
        wb.save(excel_file)
        wb.close()

        # AUTO FIT ------
        # excel = Dispatch('Excel.Application', pythoncom.CoInitialize())
        excel = CreateObject('Excel.Application', pythoncom.CoInitialize())

        wb_disptch = excel.Workbooks.Open(excel_file)
        excel.ActiveSheet.Columns.AutoFit()
        wb_disptch.Close(SaveChanges=1)
        excel.Quit()

        # ------- autofit

    def check_done(self, save_path, file_type, startswith=None):
        """
        :param save_path: lugar que checka tipo de arquivo
        :param file_type: extension
        :param startswith: same...
        :return:
        """
        from os import listdir

        for file in listdir(save_path):
            print(file)
            if file.endswith(file_type):
                if startswith is None:
                    print('CERTIFICADO EXISTENTE, NÃO PROSSEGUE')
                    return False
                elif file.startswith(startswith):
                    print('CERTIFICADO EXISTENTE, NÃO PROSSEGUE')
                    return False

        print('\033[1;35mPROSSEGUE\033[m')
        return True

    def captcha_hacking(self):
        """
        :param driver:
        :return:
        SbFConverter() -> class called

        Tremembé... rs
        """
        from pgdas_fiscal_oesk.sbfconverter import SbFConverter
        driver = self.driver

        from pyautogui import hotkey
        from pyperclip import paste
        addon = driver.find_elements(By.CLASS_NAME, 'input-group-addon')[0]
        img = addon.find_element(By.TAG_NAME, 'img')
        img_name = 'hacking.png'
        img.screenshot(img_name)
        SbFConverter(img_name)

        continua = paste()
        return continua
