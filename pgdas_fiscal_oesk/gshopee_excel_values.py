# from default import NewSetPaths, ExcelToData
from default.sets.init_email import EmailExecutor
from default.sets import InitialSetting
from win32com import client as win32
import os
from time import sleep
import pandas as pd
from openpyxl import load_workbook

# _wb_exists = self.walget_searpath
# f'{__r_social[:__r_social.find(" ")]}_{__cnpj}.xlsx'


class ShopeeExcel(InitialSetting):
    shall_sleep = True

    def __init__(self, *args, compt):
        # __r_social, __cnpj, __cpf = args
        self.compt = compt

        self.client_path = self.files_pathit('', self.compt)
        self.make_valor_total_formulas()

    @property
    def _shopee_reports_exported(self):
        emission_report = self.walget_searpath(
            'emission_report', os.path.dirname(self.client_path), 1)
        # export_report = self.walget_searpath(
        #     '[export_report]', os.path.dirname(self.client_path), 1)
        export_report = []

        reports = emission_report + export_report

        return reports

    def make_valor_total_formulas(self):
        from openpyxl.utils.formulas import FORMULAE
        import os

        for report in self._shopee_reports_exported:
            wb = load_workbook(report)
            ws = wb.active
            (ws.title)

            # Loop para percorrer a coluna Q
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=17, max_col=17):
                for cell in row:
                    # Obtém a fórmula da célula atual
                    # "SUM" in FORMULAE
                    # "VALUE" in FORMULAE

                    formula = f'=_xlfn.VALUE(SUBSTITUTE(SUBSTITUTE(VALUE($L{cell.row}), "$ ", ""), "$ ", ""))'
                    cell.value = formula

            # new_filename = os.path.splitext(report)[0] + '_modifiedds.xlsx'
            ws['Q2'].value = 'FORMULA'
            ws['R3'].value = f'=_xlfn.SUMIF(H3:H{ws.max_row},"Venda",Q3:Q{ws.max_row})'

            self._format_worksheet(ws)
            wb.save(report)
            print(f'saving report... {report}')

    def read_total_value(self):
        for report in self._shopee_reports_exported:
            wb = load_workbook(report, data_only=True)
            ws = wb.active
            total = ws['R3'].value
            # TODO colocar o valor no BD talvez?... Integrar ao Backend???

    def _format_worksheet(self, ws):
        # --------------------------------------------------------
        # Check if all cells within the range are empty
        all_empty = all(ws.cell(row=row, column=col).value is None
                        for row in range(1, 1 + 1)
                        for col in range(1, 30 + 1))
        if all_empty:
            # if empty, delete it
            ws.delete_rows(ws.min_row, 1)

        # Set the default row height (in points)
        default_row_height = 12.75

        # Loop through all rows and set the height to the default
        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = default_row_height

# =@VALOR(SUBSTITUIR(SUBSTITUIR(VALOR(L2); "$ "; ""); "$ "; ""))
# =VALOR(SUBSTITUIR(SUBSTITUIR(VALOR(L2); "$ "; ""); "$ "; ""))
