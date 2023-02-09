
# from pgdas_fiscal_oesk.silas_abre_g5_loop_v9_iss import G5
from default.webdriver_utilities.pre_drivers import pgdas_driver, pgdas_driver_ua, ginfess_driver

from default.sets import get_compt
from pgdas_fiscal_oesk import Consultar

from default.webdriver_utilities import WDShorcuts
from default.sets import InitialSetting
from pgdas_fiscal_oesk.defis_utils.legato import Legato
from pgdas_fiscal_oesk.defis_utils.legato import transformers as tfms
import os

from default.interact import *

from selenium.webdriver.common.action_chains import ActionChains

from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.common.exceptions import NoSuchElementException, ElementClickInterceptedException, UnexpectedAlertPresentException, TimeoutException


from default.webdriver_utilities.pre_drivers import pgdas_driver, pgdas_driver_ua
from time import sleep


import openpyxl


COMPT = get_compt(-1)
CONS = Consultar(COMPT)


# pdf2jpg()
# jpg2txt()

class __VisualizaTicket(InitialSetting, Legato):

    contador = 1

    def __init__(self, compt_file=None):
        """
        :param compt_file: from GUI

        # remember past_only arg from self.get_atual_competencia
        """
        import pandas as pd
        from default.webdriver_utilities.pre_drivers import pgdas_driver

        # O vencimento DAS(seja pra qual for a compt) está certo, haja vista que se trata do mes atual

        sh_name = 'DEFIS'

        compt = f"DEFIS_{self.y()}"

        # transcrevendo compt para que não confunda com PGDAS
        _path = os.path.dirname(CONS.MAIN_FILE)
        excel_file_name = os.path.join(_path,
                                       'DEFIS', f'{self.y()-1}-DEFIS-anual.xlsx')
        pdExcelFile = pd.ExcelFile(excel_file_name)
        # agora eu posso fazer downloalds sem me preocupar tendo a variável path

        msh = pdExcelFile.parse(sheet_name=str(sh_name))
        col_str_dic = {column: str for column in list(msh)}

        msh = pdExcelFile.parse(sheet_name=str(sh_name), dtype=col_str_dic)
        READ = self.le_excel_each_one(msh)
        self.after_READ = self.readnew_lista(READ, False)
        after_READ = self.after_READ

        for i, CNPJ in enumerate(after_READ['CNPJ']):
            # ####################### A INTELIGENCIA EXCEL ESTÁ SEM OS SEM MOVIMENTOS NO MOMENTO
            _cliente = after_READ['Razão Social'][i]
            _ja_declared = after_READ['Declarado'][i].upper().strip()
            _cod_sim = after_READ['Código Simples'][i]
            _cpf = after_READ['CPF'][i]
            _cert_or_login = after_READ['CERTORLOGIN'][i]

            # Dirfis exclusivos search
            _dirf_sch = after_READ['DIRF'][i]

            # self.dirf_nome = _dirf_sch if _dirf_sch != '-' else _cliente
            self.dirf_nome = CNPJ

            if _cliente == '':
                break

            if _ja_declared not in ['S', 'OK', 'FORA']:
                # ############################################################################################ ↓
                # self.client_path = self.files_pathit(_cliente, compt)
                self.client_path = self.files_pathit(
                    _cliente, compt)

                file_pdf = 'VisualizaTicket.pdf'

                dir_searched_now = self.client_path
                file_src = ''.join([os.path.join(dir_searched_now, fname) for fname in os.listdir(dir_searched_now) if
                                    fname == file_pdf])
                #  file_searched = os.path.basename(os.path.normpath(file_searched_path))
                if file_src != '':
                    now_txts = tfms.pdf2txt(file_src)
                    # Consegui finalmente essa maravilha...
                    self.here_scrap(now_txts)

    def here_scrap(self, txt):

        pasinit = txt

        user = splitxt = txt.split()
        TXT_USED = ' '.join(user)

        # input(used)

        backup1 = TXT_USED

        # ############################################## caças responsivas
        # ####### caça objeto social
        from re import finditer, findall
        # first = [(m.start(), m.end()) for m in finditer("NACIONALIDADE BRASILEIRA")]

        finder = [(m.start(), m.end())
                  for m in finditer("NACIONALIDADE BRASILEIRA", txt)]

        nomes_list = []
        for inindx, endindx in finder:
            __pnome = txt[inindx-100:inindx-1]
            __pnome = "".join(__pnome.split("\n")[-1:])
            nome = "".join(__pnome.split(",")[0])
            nomes_list.append(nome)

        # input("treste")
        # ----------------------------
        # ####### caça CPFs

        cpfs = []
        for e, cpf in enumerate(TXT_USED.split()):
            cpf = self.str_with_mask(cpf, '000.000.000-00')
            if cpf:
                cpfs.append(cpf)

        cpfs = list(dict.fromkeys(cpfs))
        # -----------------------------
        cotas = []
        # preciso fazer o set pois ele vai usar os ultimos
        for v1, v2 in zip(range(0, len(TXT_USED)), range(20, len(TXT_USED))):
            if v2 == len(TXT_USED):
                break
            if '$'.lower() in TXT_USED[v1].lower():
                val = TXT_USED[v1:v2]
                try:
                    val = val[:val.index(',')+3]

                    # cotas.add(val)
                    cotas.append(val)
                except ValueError:
                    pass

        if 'RETIRA-SE' in splitxt:
            retirado = cpfs[-1]
            cpfs = cpfs[:-1]
            # nomes = nomes[:-1]
        print(nomes_list)
        dalecotas = [f.split()[1] for f in cotas]
        dalecotas = [f.replace('.', '') for f in dalecotas]
        dalecotas = [f'R$ {r}' for r in dalecotas]
        print(cpfs)
        print(self.dirf_nome)
        # print('-'*30)
        if len(cpfs) > 3:
            cpfs = cpfs[-2:]
        for nome, cpf, cota in zip(nomes_list, cpfs, dalecotas):
            self.contador += 1
            ws2.cell(self.contador, 1).value = self.dirf_nome
            ws2.cell(self.contador, 2).value = cpf
            ws2.cell(self.contador, 3).value = nome
            ws2.cell(self.contador, 4).value = cota

        print(dalecotas)

        print('-'*30)


my_file = os.path.dirname(CONS.MAIN_FILE)
my_file = os.path.join(my_file, 'DEFIS', f'usando.xlsx')

mcp = openpyxl.load_workbook(my_file)
wks = mcp.worksheets
ws1 = wks[0]
ws2 = wks[1]
ws3 = wks[2]


def main():
    __VisualizaTicket()
    # mcp.save(os.path.dirname(InitialSetting().compt_and_filename()[1])+'\\teste.xlsx')
    mcp.save(my_file)
