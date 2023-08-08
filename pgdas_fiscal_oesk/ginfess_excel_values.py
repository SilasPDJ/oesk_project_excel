# from default import NewSetPaths, ExcelToData
from default.sets.init_email import EmailExecutor
from default.sets import InitialSetting
from win32com import client as win32
import os
from time import sleep
import pandas as pd

# _wb_exists = self.walget_searpath
# f'{__r_social[:__r_social.find(" ")]}_{__cnpj}.xlsx'


class ExcelValuesPreensh(EmailExecutor, InitialSetting):
    shall_sleep = True

    def __init__(self, *args, compt, shall_sleep=False):
        # __r_social, __cnpj, __cpf = args
        __r_social = args[0]
        self.compt = compt

        self.client_path = self.files_pathit(__r_social.strip(), self.compt)
        self.make_valor_total_formulas()

    @property
    def _shopee_reports_exported(self):
        emission_report = self.walget_searpath(
            'emission_report', os.path.dirname(self.client_path), 1)
        # export_report = self.walget_searpath(
        #     '[export_report]', os.path.dirname(self.client_path), 1)
        export_report = []

        reports = emission_report + export_report

        return reports

    def make_valor_total_formulas(self):
        from openpyxl import load_workbook
        from openpyxl.utils.formulas import FORMULAE
        from openpyxl.styles import numbers
        import os

        for report in self._shopee_reports_exported:
            wb = load_workbook(report, data_only=True)
            ws = wb.active
            (ws.title)

            # Loop para percorrer a coluna Q
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=17, max_col=17):
                for cell in row:
                    # Obtém a fórmula da célula atual
                    # "SUM" in FORMULAE
                    # "VALUE" in FORMULAE

                    formula = f'=_xlfn.VALUE(SUBSTITUTE(SUBSTITUTE(VALUE($L{cell.row}), "$ ", ""), "$ ", ""))'
                    cell.value = formula

            # new_filename = os.path.splitext(report)[0] + '_modifiedds.xlsx'
            ws['Q2'].value = 'FORMULA'
            ws['R3'].value = f'=_xlfn.SUMIF(H3:H{ws.max_row},"Venda",Q3:Q{ws.max_row})'
            wb.save(report)


# =@VALOR(SUBSTITUIR(SUBSTITUIR(VALOR(L2); "$ "; ""); "$ "; ""))
# =VALOR(SUBSTITUIR(SUBSTITUIR(VALOR(L2); "$ "; ""); "$ "; ""))
